import pandas as pd
import requests
import os

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from asreview import open_state
from langdetect import detect

from evidence import insert_evidence
from pk import process_pk

config = {}
with open('config.txt', 'r') as config_f:
    for line in config_f:
        if '=' not in line:
            continue
        key, value = line.split('=', 1)
        config[key.strip()] = value.strip()


def get_excels():
    directory = os.path.join(os.getcwd(), "ExcelFiles")

    xlsx_files = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~$'):
                xlsx_files.append(os.path.join(root, file))

    return xlsx_files


def asreview_file_exists(output_file):
    # Get filename
    filename, _ = os.path.splitext(os.path.basename(output_file))

    # Check if .asreview file exists with same name
    return os.path.exists(f'ASReviewFiles/{filename}.asreview')


def exclude_languages(df):
    print("Excluding non-English, non-Dutch and non-German articles...")

    langdetect_allowed = ['en', 'de', 'nl']
    asreview_allowed = ['eng', 'dut', 'ger', 'English']

    removed_rows = []
    for row in df.itertuples():
        # Paper contains a language value
        if pd.notna(row.language) and row.language not in asreview_allowed:
            removed_rows.append(row.Index)
        # Paper does not contain a language value (seldom)
        elif pd.isna(row.language):
            # Check if it has an original publication, and detect language
            if pd.notna(row.original_publication) and detect(row.original_publication) not in langdetect_allowed:
                removed_rows.append(row.Index)

    df = df.drop(removed_rows)
    df = df.reset_index(drop=True)

    return df


def process_doi(df):
    print("Processing doi column...")
    # Cycle through rows and transform doi into links
    for row in df.itertuples():
        if pd.notna(row.doi) and row.doi != '':
            df.at[row.Index, 'doi'] = f'https://www.doi.org/{row.doi}'

    return df


def check_full_text_availability(pubmed_id, worldcat_url):
    print(f"-Fetching online availability of article with PubMed code: {pubmed_id}")
    response = requests.get(worldcat_url + str(pubmed_id))

    # Check if the request was successful
    if response.status_code != 200:
        print(f"--Failed to retrieve the page. Status code: {response.status_code}")
        return 0
    
    # Parse the response content
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Check for indicators of full-text availability
    full_text_links = soup.find_all('section', class_='fullTextRecord')
    
    if len(full_text_links) > 0:
        print(f"--Full text access found for your institution.")
        return 1
    else:
        # Sometimes the pubmed code could be erroneous (or not a pubmed code at all) and no article is found at all
        no_results_found = soup.find_all('div', id="no-result-alert")
        if len(no_results_found) > 0:
            print(f"--Pubmed ID {pubmed_id} did not retrieve any articles. Please manually check for full-text availability!")
            return 0
        else:
            print(f"--No full text access found for your institution. Removing article.")
            return -1


def ensure_full_texts_only(df):
    # First check if it is turned on in config
    if config.get('only_full_texts', 0) == 'true' and 'institution_worldcat_url' in config:
        print("Checking for full-text capabilities...")
        worldcat_url = f'{config["institution_worldcat_url"]}/atoztitles/link?id=pmid:'
    else:
        return df
    
    dubious_rows = []
    for row in df.itertuples():
        has_full_text = check_full_text_availability(row.accession_number, worldcat_url)
        if has_full_text != 1:
            dubious_rows.append(0)
        else:
            dubious_rows.append(1)

    # Remove non-full-text papers
    df['full-text'] = dubious_rows

    return df


def add_info_from_asreview_file(file, df):
    print("Extracting labeling time and notes info from the .asreview file...")
    asreview_file, _ = os.path.splitext(os.path.basename(file))

    with open_state(f'ASReviewFiles/{asreview_file}.asreview') as state:
        asreview_df = state.get_dataset()

    # Select relevant columns and set index
    asreview_df = asreview_df[['record_id','labeling_time', 'notes']]

    # Reset index on df in order to merge
    #df.reset_index()

    # Check if 'record_id' column exists in both dataframes
    if 'record_id' not in df.columns:
        raise KeyError("'record_id' not found in df")
    if 'record_id' not in asreview_df.columns:
        raise KeyError("'record_id' not found in asreview_df")

    df = pd.merge(df, asreview_df[['record_id', 'labeling_time', 'notes']], on='record_id', how='left')

    return df


def apply_hyperlinks(processed_file, df):
    print("Applying hyperlinks...")
    workbook = load_workbook(processed_file)
    worksheet = workbook.active

    # Find the column index of the 'doi' column
    doi_col_index = None
    for col_idx, col in enumerate(worksheet.iter_cols(1, worksheet.max_column), start=1):
        if col[0].value == 'doi':
            doi_col_index = col_idx
            break

    # If 'doi' column found, adjust hyperlinks
    if doi_col_index:
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=doi_col_index)
            cell.hyperlink = cell.value
            cell.style = 'Hyperlink'
    
    workbook.save(processed_file)


def process_asreview_output():
    excel_files = get_excels()

    # Cycle through each file
    for file in excel_files:
        # ensure an .asreview file with similar name exists
        if not asreview_file_exists(file):
            print(f"Could not find .asreview file in the ASReviewFiles folder" + \
                  f"for file '{file}'. Make sure it has the exact same name!")
            continue

        df = pd.read_excel(file, header=0)

        # Eclude non-English, non-Dutch and non-German papers
        df = exclude_languages(df)

        # Ensure full-text capabilities of user
        df = ensure_full_texts_only(df)

        # Set DOI urls
        df = process_doi(df)

        # Check levels of evidence
        df = insert_evidence(df)

        # Check PK-ness of study
        df = process_pk(df)

        # Add labeling time and notes
        df = add_info_from_asreview_file(file, df)

        # Write to excel
        processed_file = f'{file.split(".")[0]}_processed.xlsx'
        df.to_excel(processed_file, index=False)

        # Turn DOI into hyperlinks
        apply_hyperlinks(processed_file, df)
        


def main():
    process_asreview_output()
    input("Press Enter to exit...")

if __name__ == "__main__":
    main()